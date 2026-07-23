from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from typing import List, Dict

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

COLUMNS = [
    {"name": "display_id",    "required": True,  "width": 18, "description": "Unique ID (e.g. DSP-001)"},
    {"name": "name",          "required": True,  "width": 35, "description": "Descriptive display name"},
    {"name": "ip_address",    "required": True,  "width": 20, "description": "IPv4 or IPv6 address"},
    {"name": "location",      "required": True,  "width": 30, "description": "Physical location"},
    {"name": "status",        "required": False, "width": 16, "description": "ACTIVE, INACTIVE, or MAINTENANCE"},
    {"name": "content_profiles", "required": False, "width": 30, "description": "Comma-separated profile names"},
]


def generate_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Displays"

    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col["name"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL if not col["required"] else REQUIRED_FILL
        cell.alignment = HEADER_ALIGNMENT
        ws.column_dimensions[get_column_letter(col_idx)].width = col["width"]

    status_col = get_column_letter(5)
    status_dv = DataValidation(
        type="list",
        formula1='"ACTIVE,INACTIVE,MAINTENANCE"',
        allow_blank=True,
    )
    status_dv.error = "Status must be ACTIVE, INACTIVE, or MAINTENANCE"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)
    status_dv.add(f"{status_col}2:{status_col}1048576")

    example_data = ["DSP-001", "Main Entrance", "192.168.1.50", "HQ, Floor 1", "ACTIVE", "Wayfinding Static,Event Schedule"]
    for col_idx, value in enumerate(example_data, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 55
    legend = [
        ("Field", "Rules"),
        ("display_id", "Required. 3-30 chars, alphanumeric. Must be unique. Example: DSP-001"),
        ("name", "Required. 2-100 characters. Descriptive display name."),
        ("ip_address", "Required. Valid IPv4 or IPv6 address."),
        ("location", "Required. 2-150 characters."),
        ("status", "Optional. Defaults to ACTIVE. Must be: ACTIVE, INACTIVE, or MAINTENANCE."),
        ("content_profiles", "Optional. Comma-separated list. Unknown profiles are auto-created."),
    ]
    for row_idx, (field, rules) in enumerate(legend, start=1):
        ws2.cell(row=row_idx, column=1, value=field).font = Font(bold=True, size=11)
        ws2.cell(row=row_idx, column=2, value=rules)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


ERROR_FONT = Font(name="Calibri", size=11)
ERROR_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
ERROR_HEADER_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
ERROR_ROW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def generate_failed_rows_excel(rows: List[Dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Failed Rows"

    columns = ["Row", "display_id", "name", "ip_address", "location", "status", "content_profiles", "Errors"]
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = ERROR_HEADER_FONT
        cell.fill = ERROR_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    for row_idx, row in enumerate(rows, start=2):
        data = row.get("data", {})
        errors = row.get("errors", [])
        ws.cell(row=row_idx, column=1, value=row.get("row_number", row_idx - 1)).font = ERROR_FONT
        ws.cell(row=row_idx, column=2, value=data.get("display_id", "")).font = ERROR_FONT
        ws.cell(row=row_idx, column=3, value=data.get("name", "")).font = ERROR_FONT
        ws.cell(row=row_idx, column=4, value=data.get("ip_address", "")).font = ERROR_FONT
        ws.cell(row=row_idx, column=5, value=data.get("location", "")).font = ERROR_FONT
        ws.cell(row=row_idx, column=6, value=data.get("status", "")).font = ERROR_FONT
        ws.cell(row=row_idx, column=7, value=", ".join(data.get("content_profiles", [])) if isinstance(data.get("content_profiles"), list) else str(data.get("content_profiles", ""))).font = ERROR_FONT
        ws.cell(row=row_idx, column=8, value="; ".join([f"{e.get('field', '')}: {e.get('message', '')}" for e in errors])).font = ERROR_FONT
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = ERROR_ROW_FILL

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
