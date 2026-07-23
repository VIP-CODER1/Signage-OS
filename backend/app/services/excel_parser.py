import re
from io import BytesIO
from typing import List, Dict, Any, Set, Tuple
from openpyxl import load_workbook
from app.models.display import Display
from app.models.content_profile import ContentProfile
from app.schemas.display import DisplayCreate
from app.schemas.excel import ValidationErrorItem, ValidatedRow, ValidationResponse, ValidationSummary

IPV4_PATTERN = re.compile(
    r"^(?:(?:25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(?:25[0-5]|2[0-4]\d|[01]?\d\d?)$"
)
IPV6_PATTERN = re.compile(
    r"^(?:(?:[0-9a-fA-F]{1,4}:){7}[0-9a-fA-F]{1,4}|"
    r"(?:[0-9a-fA-F]{1,4}:){1,7}:|"
    r"(?:[0-9a-fA-F]{1,4}:){1,6}:[0-9a-fA-F]{1,4}|"
    r"(?:[0-9a-fA-F]{1,4}:){1,5}(?::[0-9a-fA-F]{1,4}){1,2}|"
    r"(?:[0-9a-fA-F]{1,4}:){1,4}(?::[0-9a-fA-F]{1,4}){1,3}|"
    r"(?:[0-9a-fA-F]{1,4}:){1,3}(?::[0-9a-fA-F]{1,4}){1,4}|"
    r"(?:[0-9a-fA-F]{1,4}:){1,2}(?::[0-9a-fA-F]{1,4}){1,5}|"
    r"[0-9a-fA-F]{1,4}:(?:(?::[0-9a-fA-F]{1,4}){1,6})|"
    r":(?:(?::[0-9a-fA-F]{1,4}){1,7}|:)|"
    r"fe80:(?::[0-9a-fA-F]{0,4}){0,4}%[0-9a-zA-Z]{1,}|"
    r"::(?:ffff(?::0{1,4}){0,1}:){0,1}"
    r"(?:(?:25[0-5]|(?:2[0-4]|1{0,1}[0-9]){0,1}[0-9])\.){3}"
    r"(?:25[0-5]|(?:2[0-4]|1{0,1}[0-9]){0,1}[0-9])|"
    r"[0-9a-fA-F]{1,4}:(?:(?::[0-9a-fA-F]{1,4}){1,4}))$"
)

VALID_STATUSES = {"ACTIVE", "INACTIVE", "MAINTENANCE"}
REQUIRED_COLUMNS = ["display_id", "name", "ip_address", "location"]
MAX_FILE_SIZE = 10 * 1024 * 1024


class ExcelParseError(Exception):
    def __init__(self, detail: str, status_code: int = 422):
        self.detail = detail
        self.status_code = status_code


async def validate_excel(file_content: bytes, filename: str) -> dict:
    if not filename.endswith(".xlsx"):
        raise ExcelParseError("File must be a .xlsx Excel file", 422)

    if len(file_content) > MAX_FILE_SIZE:
        raise ExcelParseError("File exceeds 10MB maximum size", 422)

    try:
        wb = load_workbook(BytesIO(file_content), read_only=True)
    except Exception:
        raise ExcelParseError("File is not a valid Excel workbook", 422)

    if "Displays" not in wb.sheetnames:
        raise ExcelParseError("Missing required sheet 'Displays'", 422)

    ws = wb["Displays"]

    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if header_row is None:
        raise ExcelParseError("Empty workbook — no header row found", 422)

    missing = [col for col in REQUIRED_COLUMNS if col not in header_row]
    if missing:
        raise ExcelParseError(f"Missing column(s): {', '.join(missing)}", 422)

    column_indices = {name: idx for idx, name in enumerate(header_row)}

    existing_ids = {d.display_id async for d in Display.find_all()}
    seen_ids: Set[str] = set()

    preview = []
    valid_count = 0
    row_number = 1

    for row in ws.iter_rows(min_row=2):
        row_number += 1
        row_values = [cell.value for cell in row]

        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        data = {}
        errors = []

        for col_name, col_idx in column_indices.items():
            value = row_values[col_idx] if col_idx < len(row_values) else None
            data[col_name] = str(value).strip() if value else ""

        display_id = data.get("display_id", "")
        if not display_id:
            errors.append(ValidationErrorItem(field="display_id", message="Required field is empty"))
        elif len(display_id) < 3 or len(display_id) > 30:
            errors.append(ValidationErrorItem(field="display_id", message="Must be 3-30 characters"))
        elif not display_id.replace("-", "").replace("_", "").isalnum():
            errors.append(ValidationErrorItem(field="display_id", message="Must be alphanumeric (hyphens/underscores allowed)"))
        elif display_id in seen_ids:
            errors.append(ValidationErrorItem(field="display_id", message="Duplicate display_id encountered in same sheet"))
        elif display_id in existing_ids:
            errors.append(ValidationErrorItem(field="display_id", message="Display ID already exists in database"))

        if display_id:
            seen_ids.add(display_id)

        name_val = data.get("name", "")
        if not name_val:
            errors.append(ValidationErrorItem(field="name", message="Required field is empty"))
        elif len(name_val) < 2 or len(name_val) > 100:
            errors.append(ValidationErrorItem(field="name", message="Must be 2-100 characters"))

        ip_val = data.get("ip_address", "")
        if not ip_val:
            errors.append(ValidationErrorItem(field="ip_address", message="Required field is empty"))
        elif not (IPV4_PATTERN.match(ip_val) or IPV6_PATTERN.match(ip_val)):
            errors.append(ValidationErrorItem(field="ip_address", message="Invalid IP address format"))

        loc_val = data.get("location", "")
        if not loc_val:
            errors.append(ValidationErrorItem(field="location", message="Required field is empty"))
        elif len(loc_val) < 2 or len(loc_val) > 150:
            errors.append(ValidationErrorItem(field="location", message="Must be 2-150 characters"))

        status_val = data.get("status", "ACTIVE")
        if not status_val:
            data["status"] = "ACTIVE"
        elif status_val not in VALID_STATUSES:
            errors.append(ValidationErrorItem(field="status", message="Value must be ACTIVE, INACTIVE, or MAINTENANCE"))

        cp_val = data.get("content_profiles", "")
        cp_list = []
        if cp_val:
            cp_list = [p.strip() for p in cp_val.split(",") if p.strip()]
            for profile_name in cp_list:
                exists = await ContentProfile.find_one(ContentProfile.name == profile_name)
                if not exists:
                    errors.append(ValidationErrorItem(
                        field="content_profiles",
                        message=f"Content profile '{profile_name}' will be auto-created but has warning flag",
                    ))
        data["content_profiles"] = cp_list

        is_valid = len(errors) == 0
        if is_valid:
            valid_count += 1

        preview.append(ValidatedRow(
            row_number=row_number,
            is_valid=is_valid,
            data=data,
            errors=errors,
        ))

    total_rows = len(preview)
    return ValidationResponse(
        is_valid=(valid_count == total_rows),
        summary=ValidationSummary(
            total_rows=total_rows,
            valid_rows=valid_count,
            failed_rows=total_rows - valid_count,
        ),
        preview=preview,
    )


async def bulk_resolve_content_profiles(all_names: Set[str]) -> Dict[str, ContentProfile]:
    existing = await ContentProfile.find(
        {"name": {"$in": list(all_names)}}
    ).to_list()
    result = {p.name: p for p in existing}
    for name in all_names:
        if name and name not in result:
            profile = ContentProfile(name=name, priority="MEDIUM")
            await profile.insert()
            result[name] = profile
    return result


async def commit_validated_rows(validated_data: List[dict]) -> dict:
    all_profile_names: Set[str] = set()
    for row_data in validated_data:
        profiles = row_data.get("content_profiles", [])
        if isinstance(profiles, list):
            all_profile_names.update(profiles)
        elif isinstance(profiles, str):
            all_profile_names.update(p.strip() for p in profiles.split(",") if p.strip())

    profiles_cache = await bulk_resolve_content_profiles(all_profile_names)

    displays_to_insert = []
    commit_errors = []

    for idx, row_data in enumerate(validated_data):
        try:
            create_data = DisplayCreate(**row_data)
            profile_names = create_data.content_profiles
            resolved = [profiles_cache[name] for name in profile_names if name in profiles_cache]

            display = Display(
                display_id=create_data.display_id,
                name=create_data.name,
                ip_address=create_data.ip_address,
                location=create_data.location,
                status=create_data.status,
                content_profiles=resolved,
            )
            displays_to_insert.append(display)
        except Exception as e:
            commit_errors.append({
                "row": idx + 1,
                "error": str(e),
            })

    if displays_to_insert:
        await Display.insert_many(displays_to_insert)

    return {
        "inserted": len(displays_to_insert),
        "failed": len(commit_errors),
        "total": len(validated_data),
        "errors": commit_errors,
    }
