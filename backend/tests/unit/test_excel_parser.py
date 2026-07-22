import pytest
from io import BytesIO
from openpyxl import Workbook
from app.services.excel_parser import (
    IPV4_PATTERN, IPV6_PATTERN, VALID_STATUSES,
    validate_excel, ExcelParseError,
)


class TestIPValidation:
    @pytest.mark.parametrize("valid_ip", [
        "192.168.1.1",
        "10.0.0.1",
        "172.16.0.1",
        "255.255.255.255",
        "0.0.0.0",
    ])
    def test_valid_ipv4(self, valid_ip):
        assert IPV4_PATTERN.match(valid_ip)

    @pytest.mark.parametrize("invalid_ip", [
        "256.0.0.1",
        "192.168.1",
        "192.168.1.1.1",
        "abc.def.ghi.jkl",
        " 192.168.1.1",
        "",
    ])
    def test_invalid_ipv4(self, invalid_ip):
        assert not IPV4_PATTERN.match(invalid_ip)

    @pytest.mark.parametrize("valid_ipv6", [
        "::1",
        "2001:db8::1",
        "fe80::1",
        "2001:0db8:85a3:0000:0000:8a2e:0370:7334",
    ])
    def test_valid_ipv6(self, valid_ipv6):
        assert IPV6_PATTERN.match(valid_ipv6)


class TestStatusValidation:
    def test_valid_statuses(self):
        for status in ["ACTIVE", "INACTIVE", "MAINTENANCE"]:
            assert status in VALID_STATUSES

    def test_invalid_statuses(self):
        for status in ["active", "Active", "DISABLED", "", "PENDING"]:
            assert status not in VALID_STATUSES


class TestExcelParsing:
    def _create_valid_xlsx(self) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Displays"
        ws.append(["display_id", "name", "ip_address", "location", "status", "content_profiles"])
        ws.append(["DSP-101", "Screen 1", "192.168.1.10", "Zone A", "ACTIVE", "Test Profile"])
        ws.append(["DSP-102", "Screen 2", "192.168.1.11", "Zone B", "INACTIVE", ""])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _create_xlsx_with_errors(self) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Displays"
        ws.append(["display_id", "name", "ip_address", "location", "status", "content_profiles"])
        ws.append(["DSP-101", "A", "999.999.9.9", "X", "INVALID", "Unknown"])
        ws.append(["DSP-101", "", "not-an-ip", "", "", "Missing"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @pytest.mark.asyncio
    async def test_valid_parse(self):
        buf = self._create_valid_xlsx()
        result = await validate_excel(buf.read(), "test.xlsx")
        assert result.is_valid is True
        assert result.summary.total_rows == 2
        assert result.summary.valid_rows == 2
        assert result.summary.failed_rows == 0

    @pytest.mark.asyncio
    async def test_parse_with_errors(self):
        buf = self._create_xlsx_with_errors()
        result = await validate_excel(buf.read(), "test.xlsx")
        assert result.is_valid is False
        assert result.summary.failed_rows > 0
        assert any("ip_address" in str(e.field) for e in result.preview[0].errors)
        assert any("status" in str(e.field) for e in result.preview[0].errors)

    @pytest.mark.asyncio
    async def test_wrong_file_extension(self):
        with pytest.raises(ExcelParseError, match="must be a .xlsx"):
            await validate_excel(b"fake content", "test.csv")

    @pytest.mark.asyncio
    async def test_file_too_large(self):
        large_content = b"x" * (10 * 1024 * 1024 + 1)
        with pytest.raises(ExcelParseError, match="exceeds 10MB"):
            await validate_excel(large_content, "test.xlsx")

    @pytest.mark.asyncio
    async def test_missing_required_column(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Displays"
        ws.append(["name", "ip_address"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        with pytest.raises(ExcelParseError, match="Missing column"):
            await validate_excel(buf.read(), "test.xlsx")

    @pytest.mark.asyncio
    async def test_duplicate_intra_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Displays"
        ws.append(["display_id", "name", "ip_address", "location", "status", "content_profiles"])
        ws.append(["DSP-101", "Screen 1", "192.168.1.10", "Zone A", "ACTIVE", ""])
        ws.append(["DSP-101", "Screen 2", "192.168.1.11", "Zone B", "ACTIVE", ""])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = await validate_excel(buf.read(), "test.xlsx")
        assert result.is_valid is False
        dup_errors = [r for r in result.preview if not r.is_valid]
        assert any("duplicate" in str(e).lower() for r in dup_errors for e in r.errors)
